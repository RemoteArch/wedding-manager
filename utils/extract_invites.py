import argparse
import json
import os
import re
import secrets
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"


def _norm(s: str) -> str:
    s = str(s) if s is not None else ""
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^a-z0-9 ]+", "", s)
    return s


def _first_non_empty_str(value) -> str:
    if isinstance(value, pd.Series):
        for v in value.tolist():
            s = _first_non_empty_str(v)
            if s:
                return s
        return ""
    if value is None:
        return ""
    s = str(value).strip()
    return s


def _looks_like_header_row(row_values: List[str]) -> bool:
    joined = " | ".join(_norm(v) for v in row_values)
    return any(k in joined for k in ["nom", "invite", "invites", "table", "code"]) and len(joined) > 0


def _detect_header_row(df_raw: pd.DataFrame, max_scan_rows: int = 30) -> Optional[int]:
    scan_rows = min(max_scan_rows, len(df_raw))
    for i in range(scan_rows):
        row = ["" if pd.isna(v) else str(v) for v in df_raw.iloc[i].tolist()]
        if _looks_like_header_row(row):
            return i
    return None


def _build_df_with_header(df_raw: pd.DataFrame) -> pd.DataFrame:
    header_row = _detect_header_row(df_raw)
    if header_row is None:
        df = df_raw.copy()
        df.columns = [f"col_{i+1}" for i in range(df.shape[1])]
        return df

    header = df_raw.iloc[header_row].tolist()
    df = df_raw.iloc[header_row + 1 :].copy()
    df.columns = [str(h).strip() if h is not None else "" for h in header]
    return df


def _find_column(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    cols = list(df.columns)
    norm_map = {c: _norm(c) for c in cols}
    for cand in candidates:
        cand_n = _norm(cand)
        for c, n in norm_map.items():
            if cand_n and cand_n == n:
                return c
    for cand in candidates:
        cand_n = _norm(cand)
        for c, n in norm_map.items():
            if cand_n and cand_n in n:
                return c
    return None


def _iter_sheets(input_path: str) -> Dict[str, pd.DataFrame]:
    return pd.read_excel(input_path, sheet_name=None, header=None, engine="openpyxl")


def _generate_unique_code(used: set) -> str:
    while True:
        code = "".join(secrets.choice(ALPHABET) for _ in range(5))
        if code not in used:
            used.add(code)
            return code


def _sanitize_table(value) -> str:
    s = _first_non_empty_str(value)
    if not s:
        return ""
    s_norm = _norm(s)
    m = re.search(r"\b(\d{1,3})\b", s_norm)
    if m:
        return m.group(1)
    return s.strip()


_TABLE_HEADER_RE = re.compile(r"\btable\b\s*(?:n|n\s*o|n\s*°|no|n°)?\s*[:\-]??\s*(\d{1,3})", re.IGNORECASE)


def _parse_table_header(cell: str) -> Optional[str]:
    s = _first_non_empty_str(cell)
    if not s:
        return None
    m = _TABLE_HEADER_RE.search(s)
    if not m:
        return None
    return m.group(1)


def _looks_like_invite_name(cell: str) -> bool:
    s = _first_non_empty_str(cell)
    if not s:
        return False
    s_strip = s.strip()
    s_norm = _norm(s_strip)
    if not s_norm or s_norm == "nan":
        return False
    # Filter out known non-names / structural labels
    banned_contains = [
        "table",
        "liste",
        "invites",
        "invités",
        "invite",
        "total",
        "ok",
        "single",
        "dt",  # dtype artifacts
        "name:",
    ]
    if any(b in s_norm for b in banned_contains):
        return False
    # Avoid pure numbers
    if re.fullmatch(r"\d+", s_norm):
        return False
    # Require letters
    if not re.search(r"[a-z]", s_norm):
        return False
    # At least 3 characters
    if len(s_norm) < 3:
        return False
    return True


def _read_sheet_grid(input_path: str, sheet_name: Optional[str] = None) -> Dict[str, pd.DataFrame]:
    """Read Excel as a raw grid while honoring merged cells.

    Many invitation files are formatted as blocks using merged cells. Pandas loses merged
    information (only top-left has a value), which can reduce extraction coverage.
    """
    wb = load_workbook(filename=input_path, data_only=True)

    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
        return {sheet_name: _worksheet_to_dataframe(ws)}

    out: Dict[str, pd.DataFrame] = {}
    for name in wb.sheetnames:
        out[name] = _worksheet_to_dataframe(wb[name])
    return out


def _worksheet_to_dataframe(ws) -> pd.DataFrame:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # Map each cell coordinate in merged ranges to the top-left coordinate.
    merged_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
    for merged in ws.merged_cells.ranges:
        min_row, min_col, max_row_m, max_col_m = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        top_left = (min_row, min_col)
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                merged_map[(r, c)] = top_left

    values: List[List[str]] = []
    for r in range(1, max_row + 1):
        row_vals: List[str] = []
        for c in range(1, max_col + 1):
            tl = merged_map.get((r, c), (r, c))
            v = ws.cell(row=tl[0], column=tl[1]).value
            row_vals.append("" if v is None else str(v))
        values.append(row_vals)

    return pd.DataFrame(values)


def _extract_from_grid(df_raw: pd.DataFrame, sheet_name: str) -> List[Dict[str, str]]:
    df = df_raw.fillna("")
    nrows, ncols = df.shape

    current_table_by_col: Dict[int, str] = {}
    invites: List[Dict[str, str]] = []

    # Scan row-major; when we see a TABLE header in a column, it becomes active for that column.
    for r in range(nrows):
        for c in range(ncols):
            cell = df.iat[r, c]
            table_no = _parse_table_header(cell)
            if table_no is not None:
                current_table_by_col[c] = table_no

    # Second pass: assign invite names to nearest active table above in same column.
    current_table_by_col = {}
    for r in range(nrows):
        for c in range(ncols):
            cell = df.iat[r, c]
            table_no = _parse_table_header(cell)
            if table_no is not None:
                # If the TABLE header cell is part of a merged row-span/col-span, we might see the
                # same header repeated in adjacent columns (due to merged-cell propagation). Still,
                # we make it active for this column.
                current_table_by_col[c] = table_no
                continue

            if c not in current_table_by_col:
                continue

            if not _looks_like_invite_name(cell):
                continue

            raw = str(cell)
            # Some spreadsheets store multiple names in a single cell separated by new lines.
            parts = [p.strip() for p in re.split(r"\r?\n|;|/", raw) if p and p.strip()]
            if not parts:
                continue

            for p in parts:
                if not _looks_like_invite_name(p):
                    continue
                name = re.sub(r"\s+", " ", p).strip()
                invites.append(
                    {
                        "name": name,
                        "table": current_table_by_col[c],
                        "sheet": str(sheet_name),
                    }
                )

    # De-duplicate (same name/table can appear multiple times due to merges or repeated blocks)
    seen = set()
    unique: List[Dict[str, str]] = []
    for it in invites:
        key = (_norm(it["name"]), it["table"], it["sheet"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(it)

    return unique


def extract_invites(
    input_path: str,
    sheet: Optional[str] = None,
    overwrite_codes: bool = False,
) -> List[Dict[str, str]]:
    sheets = _read_sheet_grid(input_path, sheet_name=sheet)
    if sheet is not None:
        if not isinstance(sheets, dict):
            raise SystemExit("Unexpected sheet reading result")
        if sheet not in sheets:
            raise SystemExit(f"Sheet '{sheet}' not found. Available: {', '.join(sheets.keys())}")

    used_codes = set()
    invites: List[Dict[str, str]] = []

    if isinstance(sheets, pd.DataFrame):
        sheets = {sheet or "Sheet1": sheets}

    for sheet_name, df_raw in sheets.items():
        extracted = _extract_from_grid(df_raw, sheet_name=str(sheet_name))
        for it in extracted:
            it["code"] = _generate_unique_code(used_codes) if overwrite_codes or not it.get("code") else it["code"]
        invites.extend(extracted)

    # Always (re)generate codes uniquely for all rows
    used_codes = set()
    for it in invites:
        it["code"] = _generate_unique_code(used_codes)

    return invites


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        required=True,
        help="Path to the Excel file (.xlsx)",
    )
    parser.add_argument(
        "--out",
        default=".",
        help="Output directory (default: current directory)",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional sheet name to extract",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Generate a new code even if a code already exists in Excel",
    )

    args = parser.parse_args()

    invites = extract_invites(args.input, sheet=args.sheet, overwrite_codes=args.overwrite)
    os.makedirs(args.out, exist_ok=True)

    csv_path = os.path.join(args.out, "invites.csv")
    json_path = os.path.join(args.out, "invites.json")

    pd.DataFrame(invites).to_csv(csv_path, index=False, encoding="utf-8")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(invites, f, ensure_ascii=False, indent=2)

    print(f"Extracted {len(invites)} invites")
    print(f"CSV: {csv_path}")
    print(f"JSON: {json_path}")


if __name__ == "__main__":
    main()
